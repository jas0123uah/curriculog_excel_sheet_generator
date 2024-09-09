import openpyxl, re, os
from openpyxl.comments import Comment
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
from decouple import config
class ExcelWriter:
    import pandas as pd
    from .field import Field
    def __init__(self, concatenated_dataframe:pd.DataFrame, additional_dataframes:list[pd.DataFrame], fields:list[Field], grouping_rule:str, report_name:str ): 
        """Constructor for ExcelWriter instance. Converts dataframes to rows with openpyxl's dataframe_to_rows. concatenated_dataframe is stored as concatenated_rows. additional_dataframes are stored as additional_rows"""
        self.col_lookups = []
        self.concatenated_dataframe = concatenated_dataframe
        self.additional_dataframes = additional_dataframes
        self.report_name = report_name
        self._create_output_dir()
        self.fields = fields
        self.grouping_rule = grouping_rule
        self.workbook = openpyxl.Workbook()
        
    def _create_output_dir(self):
        report_dir = os.path.join(config('TOP_OUTPUT_DIR'), Path(self.report_name).stem)
        os.makedirs(report_dir, exist_ok=True)
        self.current_output_dir = os.path.join(report_dir, 'current',)
        self.previous_output_dir = os.path.join(report_dir, 'previous')
        os.makedirs(self.current_output_dir, exist_ok=True)
        os.makedirs(self.previous_output_dir, exist_ok=True)
    def create_workbook(self): 
        """Creates an output Excel workbook titled 'output_TIMESTAMP'.xlsx TIMESTAMP is the current timestamp when the output Excel workbook is being generated. The first sheet titled "Main" is the concatenated_rows. Additional sheets correspond to additional_rows , with the sheets having names corresponding to additional_dataframe_names."""
        self._write_sheets()
        self._create_column_lookups()
        
        for worksheet_index, worksheet in enumerate(self.workbook.worksheets):
            for row in worksheet.iter_rows(1, worksheet.max_row):
                self.set_cells_needing_comment(row, worksheet_index)
                self.set_cells_needing_embedded_url(row, worksheet_index)
                self.format_impact_cell(row, worksheet_index)
        for worksheet_index, worksheet in enumerate(self.workbook.worksheets):
            self.delete_extra_columns(worksheet=worksheet, worksheet_index=worksheet_index)
        #Current timestamp in YYYY-MM-DD-HH-MM-SS format using 12 hour clock
        now = datetime.now().strftime('%Y_%m_%d_%I_%M_%p')
        # Move all of the xlsx files in current_output_dir to the previous_output_dir
        for file in os.listdir(self.current_output_dir):
            os.rename(os.path.join(self.current_output_dir, file), os.path.join(self.previous_output_dir, file))
        self.workbook.save(os.path.join(f'{self.current_output_dir}', f'{now}.xlsx'))
           
    
    def _write_sheets(self):
        """Write the dataframes to the output workbook."""
        # Write concatenated dataframe to sheet "Main"
        dfs = [{'Main': self.concatenated_dataframe }, *self.additional_dataframes]
        for idx, df_data in enumerate(dfs):
            for df_name, df in df_data.items():
                df_name = re.sub(INVALID_TITLE_REGEX, '_', df_name)
                if len(df_name) > 31:
                    df_name = df_name[:31]
                if df_name == '':
                    df_name = 'EMPTY'
                sheet = self.workbook.create_sheet(df_name, ) if idx != 0 else self.workbook.active
                sheet.title = df_name
                for r in dataframe_to_rows(df, index=True, header=True):
                    sheet.append(r)
                # HACK: index true and header true together cause a blank row to be added at the top of the sheet. 
                sheet.delete_rows(2)
                # Freeze the top row of the additional sheet
                sheet.freeze_panes = 'A2'

    def set_cells_needing_comment(self, row, worksheet_index): 
        """Given an input row use column_names_needing_comments to find which cells in a row need a comment and give them a comment."""
        ##Loop over a dict mapping columns that need comments (target columns) to (comment_columns)
        worksheet_col_lookup = self.col_lookups[worksheet_index]
        for target_column, target_column_info in worksheet_col_lookup.items():
            ##Get cell in row needing comment.
            cell_needing_comment = self.find_cell_by_column(row=row, column_name=target_column,  worksheet_index=worksheet_index)
            ##Find other column which contains the comment text
            if target_column_info["comment_field"] != '':
                comment_column = target_column_info['comment_field']
                cell_with_comment = self.find_cell_by_column(row=row, column_name=comment_column, worksheet_index=worksheet_index)
                ##Call set_cell_comment IF CHECK IGNORES THE HEADER
                if cell_with_comment.value != comment_column:
                    self.set_cell_comment(cell_needing_comment, cell_with_comment.value)
    def set_cells_needing_embedded_url(self, row, worksheet_index):
        """Given an input row use columns_with_embedded_urls to find which cells in a row need the proposal url embedded and embed the url.""" 
        col_lookup = self.col_lookups[worksheet_index]
        columns_needing_embedded_url = [target_col for target_col, target_col_data in col_lookup.items() if col_lookup[target_col]['embed_url'] is True]
        for column_name in columns_needing_embedded_url:
            cell_needing_embedded_url = self.find_cell_by_column(row, column_name, worksheet_index)
            cell_with_embedded_url = self.find_cell_by_column(row, 'URL', worksheet_index)
            url = cell_with_embedded_url.value
            if cell_needing_embedded_url.value!= column_name:
                cell_needing_embedded_url.hyperlink = url
                cell_needing_embedded_url.style = 'Hyperlink'
    def find_cell_by_column(self, row, column_name:str, worksheet_index:int = 0): 
        """Given an input row and column_name, find the cell corresponding to column_name and return its value."""
        col_lookup = self.col_lookups[worksheet_index]
        idx = col_lookup[column_name]["idx"]
        
        target_cell = list(filter(lambda cell: cell.column == idx +1, row))[0]
        return target_cell

    def set_cell_comment(self, cell, comment_text:str): 
        """Set a given cell's comment."""
        comment = Comment(comment_text, 'OpenPyxl')
        cell.comment = comment
    def _create_column_lookups(self):
        """Create a lookup of column names and indices for the given column name for each worksheet in the workbook. Lookup used for manipulating data in each worksheet."""
        #https://stackoverflow.com/questions/51478413/select-a-column-by-its-name-openpyxl
        self.col_lookups = []
        column_idx  = 0
        # Loop over the worksheets in the workbook
        col_lookup = {}
        for sheet in self.workbook.worksheets:
            # Loop over the columns in the worksheet
            for col in sheet.iter_cols(1, sheet.max_column):
                column_name = col[0].value
                field = next((field for field in self.fields if field.field_name == column_name), None)
                comment_field = field.comment_field if field is not None else ''
                dont_return_field = field.dont_return_field if field is not None else ''
                embed_url = field.embed_url if field is not None else False

                col_lookup[column_name] = {'idx': column_idx, 'comment_field': comment_field, 'dont_return_field': dont_return_field, 'embed_url': embed_url}
                column_idx += 1
            self.col_lookups.append(col_lookup)
            column_idx = 0
            col_lookup = {}
    def delete_extra_columns(self, worksheet, worksheet_index): 
        """Delete columns from rows which were used solely for grouping OR contain comment text as a value. (Comments are stored as fields in the pandas dataframe this step removes those columns)."""
        column_lookup = self.col_lookups[worksheet_index]
        columns_to_delete = [column_lookup[col]['comment_field'] for col in column_lookup if column_lookup[col]['comment_field'] != '']
        

        #Loop over column_names_needing_comments
        for col_name in columns_to_delete:
            col_idx = column_lookup[col_name]['idx']
            worksheet.delete_cols(col_idx+1, 1)
    def delete_group_by_col_name(self,worksheet):
        """Looks for a column used to create additional sheets in the output excel workbook. If that field was not requested, that column is deleted."""
        requested_columns = list(map(lambda field: field.field_name, self.fields))
        if self.grouping_rule and self.grouping_rule not in requested_columns:
            col_idx = self.col_lookup[self.grouping_rule]
            worksheet.delete_cols(col_idx, 1) 
    
    def get_impact_color(self, impact):
        """Returns the color the cell should be based on its impact"""
        red_fill = PatternFill(start_color='FFC7CE',
                   end_color='FFC7CE',
                   fill_type='solid')
        gold_fill = PatternFill(start_color='FFEB9C',
                   end_color='FFEB9C',
                   fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCC',
                   end_color='C6EFCC',
                   fill_type='solid')
        
        if impact == 'High':
            return red_fill
        elif impact == 'Mid':
            return gold_fill
            
        elif impact == 'Low':
            return green_fill
        else:
            return None
    
    def get_text_color(self, impact):
        if impact == 'High':
            return Color(rgb='9C0006')
        elif impact == 'Mid':
            return Color(rgb='9C5700')
        elif impact == 'Low':
            return Color(rgb='006100')
    def format_impact_cell(self, row, worksheet_index=0):
        """Function for formatting the impact column cells with proper background and font color"""
        cell = self.find_cell_by_column(row=row, column_name='Impact Level', worksheet_index=worksheet_index)
        cell_color = self.get_impact_color(cell.value)
        text_color = self.get_text_color(cell.value)
        if cell_color != None:
            cell.fill = cell_color #https://stackoverflow.com/questions/30484220/fill-cells-with-colors-using-openpyxl
            cell.font = Font(color=text_color)
