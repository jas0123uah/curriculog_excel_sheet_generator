from .field import Field
from .filter import Filter
from .sorting_rule import SortingRule
from pprint import pprint
import datetime
import logging, sys
# Configure root logger  
logging.basicConfig(format='%(asctime)s | %(name)s | %(levelname)s |', 
                    filename='log.txt',
                    filemode='a',
                    level=logging.INFO)

logger = logging.getLogger(__name__)


# Log to file
file_handler = logging.FileHandler('log.txt')  
logger.addHandler(file_handler)

# Log to console
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
logger.addHandler(console_handler)




class ExcelInputParser:
    def __init__(self, input_workbook): 
        """Constructor for ExcelInputParser instance. Accepts the path to an Excel Workbook"""
        import openpyxl
        self.workbook = openpyxl.load_workbook(input_workbook)
        self.columns = {}
        self.fields = []
        self.filters = []
        self.sorting_rules = []
        #Default to an empty field instance
        self.grouping_rule = None
        self.api_filters = {}
        self.api_field_name_from_user_friendly_field_name = {
            'Completed Date': 'completed_date',
            'Completed Date After': 'completed_date_after',
            'Completed Date Before': 'completed_date_before',
            
            'Launch Date After': 'launch_date_after',
            'Launch Date Before': 'launch_date_before',
            'Proposed By': 'originator_id',
            'Proposal ID': 'proposal_id',
            'Proposal Name': 'proposal_name',
            'Proposal Status': 'proposed_status',
            'Proposal Type': 'proposal_type',
            'Step Name': 'step_name',
            'Current Step Started (Date)': 'step_start_date',
            'Current Step Started Before (Date)': 'step_start_date_before',
            'Current Step Started After (Date)': 'step_start_date_after',
            'Current Step Status': 'step_status',
            'Proposal Submitter First Name': 'first_name',
            'Proposal Submitter Last Name': 'last_name',
            'Proposal Submitter Email': 'email',
            'Is Remote': 'is_remote',
            'Proposal Launch Date': 'launch_date' 
        }

    def parse_workbook(self): 
        """Primary function to parse input Excel workbook. Makes use of helper functions _get_requested_fields, _get_field_filters, _get_field_comment to parse input excel workbook."""
        self._get_worksheet_by_columns()
        self._get_requested_fields()
        self._get_sorting_rules()
        self.get_grouping_rule()

    def _get_requested_fields(self): 
        """Parses the Field column in the Excel workbook to identify which proposal fields are needed from the Curriculog API. Returns an array/list of Field instances which contain filtering information specific to the Field for passing to the PandasHelper constructor ."""
        for row_cells in self.workbook.active.iter_rows(min_row=2, max_row=self.workbook.active.max_row):
            field_filter = self._get_field_filters(row_cells)
            dont_return_field = self._get_value_in_column(row=row_cells, column="Don't Return Field")
            embed_url = self._get_value_in_column(row=row_cells, column="Embed URL")
            field = Field(
                field_name=self._get_value_in_column(row=row_cells, column='Field'), 
                comment_field=self._get_value_in_column(row=row_cells, column='Field Comment'), 
                filters=field_filter,
                dont_return_field= dont_return_field if dont_return_field else False, 
                embed_url= embed_url if embed_url else False
                )
            # print('APPENDING:\n')
            # if(field.filters):
            #     pprint(vars(field.filters))
            self.fields.append(field)

    def _get_field_filters(self, row) -> Filter: 
        """For a given row in the Field column in the Excel workbook, identify the Filter for that field."""
        cols = ['Field', 'Operator', 'Value(s)']
        vals = []
        for col in cols:
            val = self._get_value_in_column(row=row, column=col)
            vals.append(val)
        field_filter = None
        if vals[1]: ##if there is no operator assume no filter.
            if type(vals[2]) is not datetime.datetime:
                values=vals[2].split(",")
            else:
                values=vals[2].strftime('%Y-%m-%d')
            field_filter = Filter(field_name=vals[0], operator=vals[1], values=values)
            #Keep the filter by itself for easy access later
            self.filters.append(field_filter)
        return field_filter


    def _get_sorting_rules(self): 
        """Parses the Sort By, Sort Order, and Custom Sort columns to create an array/list of SortingRules. Returns a list of SortingRules to pass to the PandasHelper constructor."""
        for row_cells in self.workbook.active.iter_rows(min_row=2, max_row=self.workbook.active.max_row):
            field_name = self._get_value_in_column(row=row_cells, column ='Sort By')
            #If there is no value in the column 'Sort By', assume no SortingRule for the row
            if field_name:
                sorting_rule = SortingRule(
                    field_name= field_name,
                    sort_order= self._get_value_in_column(row=row_cells, column ='Sort Order'),
                    values= self._get_value_in_column(row=row_cells, column ='Custom Sort', delimiter=","),       
                )
                self.sorting_rules.append(sorting_rule)
    def get_api_filters(self):
        """Loops over self.fields to get all fields that may be used in API filtering."""
        api_filters = {}
        #print(self.fields)
        #pprint(vars(self.fields))
        for field in self.fields:
            api_filter_field = self.api_field_name_from_user_friendly_field_name.get(field.field_name)
            ##The only operator the api supports is =. Everything else we will do on our 
            if api_filter_field and field.filters and field.filters.operator == '=':
                api_filters[api_filter_field] = field.filters.values
        self.api_filters = api_filters
        logging.info(msg=f'THE API FILTERS {self.api_filters}')
        

    def get_grouping_rule(self): 
        """Parses the Separate Sheets By column to get the field which should be used for creating additional_dataframes in the PandasHelper instance."""
        self.grouping_rule = self.workbook.active.cell(row = 2, column = self.workbook.active.max_column).value
    def _get_worksheet_by_columns(self):
        """Creates a dict of values in a given column of the active Excel spreadsheet"""
        #https://stackoverflow.com/questions/51478413/select-a-column-by-its-name-openpyxl
        col_names = {}
        Current  = 0
        for COL in self.workbook.active.iter_cols(1, self.workbook.active.max_column):
            col_names[COL[0].value] = Current
            Current += 1
        self.columns = col_names
    def _get_value_in_column(self, row, column:str, delimiter=None) -> str:
        """Given a row and column name return the value in column"""
        data = row[self.columns[column]].value
        if data is None:
            return ''
        if delimiter is None:
            return data
        
        data = data.split(delimiter)
        #Remove spaces the user likely put between words in delimiter 
        data = list(map(lambda word: word.strip(), data))
        #Return values of interest as comma separated string. No spaces.
        return ",".join(data)
